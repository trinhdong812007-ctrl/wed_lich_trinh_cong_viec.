# -*- coding: utf-8 -*-
"""
app/routes/schedule.py
Routes phân công & lịch làm việc: nhập TKB tuần, xóa lịch, tải template, nhập dữ liệu, trang động.
"""

import csv
import io
import os

import openpyxl
from datetime import timedelta, date, datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_from_directory, Response
from flask_login import login_required, current_user
from openpyxl import load_workbook

from config import BASE_DIR
from app.extensions import db
from app.models import Employee, Task, Schedule, WeekSchedule, Page, TKBFile
from app.services.helpers import normalize_text, get_week_start, touch_last_update, parse_date_value
from app.services.validation import validate_employee_payload, validate_task_payload

bp = Blueprint("schedule", __name__)


@bp.route("/download-template/<string:template_type>")
@login_required
def download_template(template_type):
    if template_type == "employees":
        filename = "employees_sample.xlsx"
    elif template_type == "tasks":
        filename = "tasks_sample.xlsx"
    elif template_type == "tkb_tuan":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TKB_tuan"
        thu_headers = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
        ws.cell(row=1, column=1, value="Ca")
        for i, thu in enumerate(thu_headers):
            ws.cell(row=1, column=i + 2, value=thu)
        ws.cell(row=2, column=1, value="Sáng")
        ws.cell(row=7, column=1, value="Chiều")
        for col in range(1, 9):
            ws.cell(row=2, column=col).font = openpyxl.styles.Font(bold=True)
            ws.cell(row=7, column=col).font = openpyxl.styles.Font(bold=True)
        for row in range(3, 7):
            for col in range(2, 9):
                ws.cell(row=row, column=col).border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin'),
                )
        for row in range(8, 12):
            for col in range(2, 9):
                ws.cell(row=row, column=col).border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin'),
                )
        ws.column_dimensions['A'].width = 10
        for col in range(2, 9):
            ws.column_dimensions[chr(64 + col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=TKB_tuan_mau.xlsx"}
        )
    else:
        return "Template không tồn tại", 404
    return send_from_directory(BASE_DIR, filename, as_attachment=True)


@bp.route("/upload-schedule", methods=["GET", "POST"])
@login_required
def upload_schedule():
    imported_count = 0
    skipped_count = 0
    errors_list = []
    today = date.today()
    week_start = get_week_start(today)
    kept_file = TKBFile.query.filter_by(week_start=week_start).first()

    if request.method == "POST":
        uploaded_file = request.files.get("file")
        if not uploaded_file or not uploaded_file.filename:
            flash("Vui lòng chọn file Excel để tải lên.", "danger")
            return render_template("schedule/upload_schedule.html", imported_count=0, skipped_count=0, errors_list=[], week_start=week_start, kept_file=kept_file)

        filename = uploaded_file.filename
        content = uploaded_file.read()
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            flash(f"Không thể đọc file: {str(e)}", "danger")
            return render_template("schedule/upload_schedule.html", imported_count=0, skipped_count=0, errors_list=[], week_start=week_start, kept_file=kept_file)

        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows or len(raw_rows) < 2:
            flash("File không có dữ liệu hợp lệ.", "danger")
            return render_template("schedule/upload_schedule.html", imported_count=0, skipped_count=0, errors_list=[], week_start=week_start, kept_file=kept_file)

        thu_map = {
            "thứ 2": 0, "thứ hai": 0, "t2": 0,
            "thứ 3": 1, "thứ ba": 1, "t3": 1,
            "thứ 4": 2, "thứ tư": 2, "t4": 2,
            "thứ 5": 3, "thứ năm": 3, "t5": 3,
            "thứ 6": 4, "thứ sáu": 4, "t6": 4,
            "thứ 7": 5, "thứ bảy": 5, "t7": 5,
            "chủ nhật": 6, "cn": 6,
        }

        headers = []
        for cell in raw_rows[0]:
            h = normalize_text(cell).lower() if cell else ""
            headers.append(h)

        col_day_map = {}
        for idx, h in enumerate(headers):
            if idx == 0:
                continue
            for key, day_idx in thu_map.items():
                if key in h:
                    col_day_map[idx] = day_idx
                    break

        # File được tải lên là nguồn dữ liệu chính thức của tuần: luôn thay thế
        # toàn bộ lịch tuần hiện tại bằng dữ liệu trong file (đồng thời giữ file
        # này làm "file đang áp dụng").
        WeekSchedule.query.filter(
            WeekSchedule.week_start == week_start
        ).delete()
        db.session.flush()

        vi_tri_sang = 0
        vi_tri_chieu = 0
        current_ca = None

        for row_idx, row in enumerate(raw_rows[1:], start=2):
            ca_cell = normalize_text(row[0]).lower() if row and row[0] else ""
            if "sáng" in ca_cell or "sang" in ca_cell:
                current_ca = "Sáng"
                vi_tri_sang = 0
                continue
            elif "chiều" in ca_cell or "chieu" in ca_cell:
                current_ca = "Chiều"
                vi_tri_chieu = 0
                continue
            elif ca_cell in ("", None) and current_ca is None:
                continue

            if current_ca is None:
                continue

            if current_ca == "Sáng":
                vi_tri_sang += 1
                vi_tri = vi_tri_sang
            else:
                vi_tri_chieu += 1
                vi_tri = vi_tri_chieu

            for col_idx, cell_value in enumerate(row):
                if col_idx == 0:
                    continue
                if col_idx not in col_day_map:
                    continue
                ma_cv = normalize_text(cell_value) if cell_value else ""
                if not ma_cv:
                    continue

                day_offset = col_day_map[col_idx]
                ngay_lam_viec = week_start + timedelta(days=day_offset)

                task = Task.query.filter_by(ma_cv=ma_cv).first()
                if not task:
                    skipped_count += 1
                    errors_list.append(f"Dòng {row_idx}: Mã CV '{ma_cv}' không tồn tại.")
                    continue

                existing = WeekSchedule.query.filter_by(
                    task_id=task.id,
                    ngay_lam_viec=ngay_lam_viec,
                    ca=current_ca
                ).first()
                if existing:
                    existing.vi_tri = vi_tri
                    existing.updated_by_id = current_user.id
                else:
                    ws = WeekSchedule(
                        task_id=task.id,
                        ngay_lam_viec=ngay_lam_viec,
                        ca=current_ca,
                        vi_tri=vi_tri,
                        week_start=week_start,
                        created_by_id=current_user.id,
                        updated_by_id=current_user.id
                    )
                    db.session.add(ws)
                imported_count += 1

        db.session.commit()
        touch_last_update()

        # Lưu/ghi đè file đang áp dụng cho tuần này vào CSDL.
        stored_file = TKBFile.query.filter_by(week_start=week_start).first()
        stored_at = datetime.utcnow()
        if stored_file:
            stored_file.filename = filename
            stored_file.content = content
            stored_file.uploaded_at = stored_at
            stored_file.created_by_id = current_user.id
        else:
            db.session.add(TKBFile(
                filename=filename,
                content=content,
                week_start=week_start,
                uploaded_at=stored_at,
                created_by_id=current_user.id
            ))
        db.session.commit()
        kept_file = TKBFile.query.filter_by(week_start=week_start).first()

        if imported_count:
            flash(f"Đã nhập {imported_count} lịch công việc cho tuần {week_start.strftime('%d/%m/%Y')}.", "success")
        if skipped_count:
            flash(f"Bỏ qua {skipped_count} ô lỗi.", "warning")

    return render_template(
        "schedule/upload_schedule.html",
        imported_count=imported_count,
        skipped_count=skipped_count,
        errors_list=errors_list,
        week_start=week_start,
        kept_file=kept_file,
    )


@bp.route("/tkb-file/download")
@login_required
def tkb_file_download():
    week_start = get_week_start(date.today())
    f = TKBFile.query.filter_by(week_start=week_start).first()
    if not f:
        flash("Chưa có file thời khóa biểu nào được áp dụng cho tuần này.", "warning")
        return redirect(url_for("schedule.upload_schedule"))
    return Response(
        f.content,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"inline; filename={f.filename}"}
    )


@bp.route("/schedule/delete-slot", methods=["POST"])
@login_required
def schedule_delete_slot():
    task = Task.query.get(request.form.get("task_id", type=int))
    ngay = parse_date_value(request.form.get("ngay_lam_viec"))
    ca = request.form.get("ca")

    if not task or not ngay or not ca:
        flash("Thiếu thông tin để xóa ô lịch.", "danger")
        return redirect(request.referrer or url_for("dashboard.lich_trinh"))

    Schedule.query.filter_by(task_id=task.id, ngay_lam_viec=ngay, ca=ca).delete()
    WeekSchedule.query.filter_by(task_id=task.id, ngay_lam_viec=ngay, ca=ca).delete()
    db.session.commit()
    touch_last_update()

    flash(f"Đã xóa công việc '{task.ten_cv}' khỏi {ngay.strftime('%d/%m/%Y')} ca {ca}.", "success")
    return redirect(request.referrer or url_for("dashboard.lich_trinh"))


@bp.route("/schedule/delete/<int:sch_id>", methods=["POST"])
@login_required
def schedule_delete(sch_id):
    sch = Schedule.query.get_or_404(sch_id)
    db.session.delete(sch)
    db.session.commit()
    touch_last_update()
    flash("Đã xóa lịch phân công.", "success")
    return redirect(request.referrer or url_for("dashboard.lich_trinh"))


@bp.route("/import-data", methods=["GET", "POST"])
@login_required
def import_data_page():
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    summary = []

    if request.method == "POST":
        import_type = request.form.get("import_type", "employees")
        uploaded_file = request.files.get("file")
        raw_data = request.form.get("raw_data", "").strip()

        if not uploaded_file and not raw_data:
            flash("Vui lòng chọn tập tin hoặc dán dữ liệu để nhập.", "danger")
            return redirect(url_for("schedule.import_data_page"))

        rows = []
        if uploaded_file and uploaded_file.filename:
            filename = uploaded_file.filename.lower()
            content = uploaded_file.read()
            if filename.endswith(".csv"):
                text_data = content.decode("utf-8-sig")
                rows = list(csv.DictReader(io.StringIO(text_data)))
            elif filename.endswith(".xlsx"):
                workbook = load_workbook(io.BytesIO(content), data_only=True)
                sheet = workbook.active
                raw_rows = list(sheet.iter_rows(values_only=True))
                if not raw_rows:
                    flash("Tập tin không chứa dữ liệu.", "danger")
                    return redirect(url_for("schedule.import_data_page"))
                headers = [normalize_text(h) for h in raw_rows[0]]
                for raw_row in raw_rows[1:]:
                    if not any(v is not None and normalize_text(v) for v in raw_row):
                        continue
                    record = {}
                    for index, header in enumerate(headers):
                        value = raw_row[index] if index < len(raw_row) else ""
                        record[header] = value
                    rows.append(record)
            else:
                flash("Chỉ hỗ trợ file .csv hoặc .xlsx.", "danger")
                return redirect(url_for("schedule.import_data_page"))
        elif raw_data:
            reader = csv.DictReader(io.StringIO(raw_data))
            if not reader.fieldnames or all(not normalize_text(fn) for fn in reader.fieldnames):
                flash("Dữ liệu CSV dán vào cần có hàng tiêu đề hợp lệ (ví dụ: ma_nv, ho_ten, bo_phan).", "danger")
                return redirect(url_for("schedule.import_data_page"))
            rows = [row for row in reader if any(normalize_text(v) for v in row.values())]

        if not rows:
            flash("Không tìm thấy dữ liệu hợp lệ trong file hoặc đoạn văn bản đã nhập.", "danger")
            return redirect(url_for("schedule.import_data_page"))

        if import_type == "employees":
            for row in rows:
                payload = {
                    "ma_nv": normalize_text(row.get("ma_nv") or row.get("Mã nhân viên") or row.get("maNV") or row.get("id")),
                    "ho_ten": normalize_text(row.get("ho_ten") or row.get("Họ tên") or row.get("Họ và tên") or row.get("hoTen") or row.get("name")),
                    "email": normalize_text(row.get("email") or row.get("Email") or row.get("mail")),
                    "bo_phan": normalize_text(row.get("bo_phan") or row.get("Bộ phận") or row.get("department")),
                    "vi_tri": normalize_text(row.get("vi_tri") or row.get("Vị trí") or row.get("position") or ""),
                }

                errors = validate_employee_payload(payload)
                if errors:
                    skipped_count += 1
                    summary.append({"row": row, "errors": errors})
                    continue

                existing_emp = Employee.query.filter_by(ma_nv=payload["ma_nv"]).first()

                if existing_emp:
                    existing_emp.ho_ten = payload["ho_ten"]
                    existing_emp.email = payload["email"]
                    existing_emp.bo_phan = payload["bo_phan"]
                    existing_emp.vi_tri = payload["vi_tri"]
                    existing_emp.updated_by_id = current_user.id
                    updated_count += 1
                else:
                    employee = Employee(
                        **payload,
                        created_by_id=current_user.id,
                        updated_by_id=current_user.id
                    )
                    db.session.add(employee)
                    imported_count += 1

            db.session.commit()
            flash(
                f"Xử lý hoàn tất! Thêm mới: {imported_count} nhân viên, Cập nhật: {updated_count} nhân viên, Bỏ qua: {skipped_count} dòng lỗi.",
                "success"
            )

        else:
            for row in rows:
                payload = {
                    "ma_cv": normalize_text(row.get("ma_cv") or row.get("Mã công việc") or row.get("maCV") or row.get("id")),
                    "ten_cv": normalize_text(row.get("ten_cv") or row.get("Tên công việc") or row.get("tenCV") or row.get("title")),
                    "ghi_chu": normalize_text(row.get("ghi_chu") or row.get("Ghi chú") or row.get("note") or row.get("description")),
                    "do_uu_tien": normalize_text(row.get("do_uu_tien") or row.get("Độ ưu tiên") or row.get("priority")) or "Trung bình",
                    "bo_phan": normalize_text(row.get("bo_phan") or row.get("Bộ phận") or row.get("department")),
                    "so_luong_nv": normalize_text(row.get("so_luong_nv") or row.get("Số lượng NV") or row.get("quantity") or "1"),
                    "thoi_luong": normalize_text(row.get("thoi_luong") or row.get("Thời lượng") or row.get("duration") or "1"),
                }

                errors = validate_task_payload(payload)
                if errors:
                    skipped_count += 1
                    summary.append({"row": row, "errors": errors})
                    continue

                existing_task = Task.query.filter_by(ma_cv=payload["ma_cv"]).first()

                if existing_task:
                    existing_task.ten_cv = payload["ten_cv"]
                    existing_task.ghi_chu = payload["ghi_chu"]
                    existing_task.do_uu_tien = payload["do_uu_tien"]
                    existing_task.bo_phan = payload["bo_phan"]
                    existing_task.so_luong_nv = int(payload["so_luong_nv"])
                    existing_task.thoi_luong = float(payload["thoi_luong"])
                    existing_task.updated_by_id = current_user.id
                    existing_task.ngay_gio = None
                    existing_task.ca_requirement = None
                    updated_count += 1
                else:
                    task = Task(
                        ma_cv=payload["ma_cv"],
                        ten_cv=payload["ten_cv"],
                        ghi_chu=payload["ghi_chu"],
                        do_uu_tien=payload["do_uu_tien"],
                        ngay_gio=None,
                        ca_requirement=None,
                        bo_phan=payload["bo_phan"],
                        so_luong_nv=int(payload["so_luong_nv"]),
                        thoi_luong=float(payload["thoi_luong"]),
                        created_by_id=current_user.id,
                        updated_by_id=current_user.id
                    )
                    db.session.add(task)
                    imported_count += 1

            db.session.commit()
            flash(
                f"Xử lý hoàn tất! Thêm mới: {imported_count} công việc, Cập nhật: {updated_count} công việc, Bỏ qua: {skipped_count} dòng lỗi.",
                "success"
            )

        return render_template(
            "schedule/import_data.html",
            import_type=import_type,
            imported_count=imported_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            summary=summary
        )

    return render_template(
        "schedule/import_data.html",
        import_type="employees",
        imported_count=0,
        updated_count=0,
        skipped_count=0,
        summary=[]
    )


@bp.route("/p/<path:slug>")
def dynamic_page(slug):
    page = Page.query.filter_by(slug=slug).first()
    if not page:
        return "Trang không tồn tại (404)", 404
    return render_template("dashboard/dynamic_page.html", page=page)


@bp.route("/pages/add", methods=["POST"])
@login_required
def page_add():
    slug = request.form.get("slug", "").strip()
    tieu_de = request.form.get("tieu_de", "").strip()
    noi_dung = request.form.get("noi_dung", "").strip()
    if not slug or not tieu_de:
        flash("Slug và Tiêu đề là bắt buộc.", "danger")
        return redirect(url_for("dashboard.lich_trinh"))
    if Page.query.filter_by(slug=slug).first():
        flash(f"Đường dẫn '/p/{slug}' đã tồn tại!", "danger")
        return redirect(url_for("dashboard.lich_trinh"))
    new_page = Page(slug=slug, tieu_de=tieu_de, noi_dung=noi_dung)
    db.session.add(new_page)
    db.session.commit()
    flash(f"Đã tạo đường dẫn mới: /p/{slug}", "success")
    return redirect(url_for("schedule.dynamic_page", slug=slug))